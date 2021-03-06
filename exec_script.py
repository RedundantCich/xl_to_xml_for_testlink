import xml.etree.cElementTree as ET
from openpyxl import load_workbook
from tools import etree_actions, file_operations
from config import reader_config as cfg


def xl_to_xml_for_testlink(columns_in_use, rows_to_skip, folder_xl_file_list):
    for file_name in folder_xl_file_list:
        workbook = load_workbook(file_name)
        sheet = workbook.active
        iter_rows = sheet.iter_rows()
        file_operations.skip_start_rows(iter_rows, rows_to_skip)

        main_test_suite = None
        number_of_current_test_suite = 1

        for row in iter_rows:
            for column in row:
                cell_value = column.value
                if cell_value is not None:
                    column_letter = column.column_letter
                    if column_letter in columns_in_use:
                        column_name = columns_in_use[column_letter][0]
                        xml_designation = columns_in_use[column_letter][1]
                        additional_action = columns_in_use[column_letter][2]
                    else:
                        continue

                    if xml_designation == 'main_parent':
                        if main_test_suite is not None:
                            number_of_current_test_suite = \
                                file_operations.write_tree_to_xml_file(file_name, main_test_suite, number_of_current_test_suite)
                        main_test_suite = etree_actions.create_main_parent(cell_value, column_name)
                        continue
                    elif xml_designation == 'sub_parent' and additional_action == 'optional_test_suite_present':
                        sub_test_suite = etree_actions.create_name_sub_parent(cell_value, column_name, main_test_suite)
                        parent_of_test_case = sub_test_suite
                        continue
                    elif xml_designation == 'sub_parent' and additional_action == 'no_additional_actions':
                        if sub_test_suite is None:
                            parent_of_test_case = main_test_suite
                        test_case = etree_actions.create_name_sub_parent(cell_value, column_name, parent_of_test_case)
                        continue
                    elif xml_designation == 'text_child':
                        etree_actions.create_text_child(cell_value, column_name, test_case)

                    if additional_action == 'start_steps':
                        steps = etree_actions.create_simple_child(test_case, 'steps')
                        case_step = 1
                    elif additional_action == 'start_single_step':
                        step = etree_actions.create_simple_child(steps, 'step')
                        etree_actions.create_text_child(case_step, 'step_number', step)
                        etree_actions.create_text_child(cell_value, column_name, step)
                    elif additional_action == 'end_single_step':
                        etree_actions.create_text_child(cell_value, column_name, step)
                        case_step += 1

        number_of_current_test_suite = \
            file_operations.write_tree_to_xml_file(file_name, main_test_suite, number_of_current_test_suite)


columns_in_use = cfg.columns_in_use
rows_to_skip = cfg.rows_to_skip
folder_xl_file_list = file_operations.list_files()

xl_to_xml_for_testlink(columns_in_use, rows_to_skip, folder_xl_file_list)
